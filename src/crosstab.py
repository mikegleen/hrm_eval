# -*- coding: utf-8 -*-
"""
Input is a CSV file produced by SurveyMonkey. The creation method is:
1. Click on "Analyze Results".
2. Step deleted.
3. Click on "Export All"
4. Select the dropdown "All response data".
5. Select the option ".XLS" (the leftmost option).
6. In "Data View", select "Original View".
7. In "Columns", select "Expanded".
8. In "Cells", select "Actual Answer Text".
9. Rename the file in the format yyyy-mm-dd_n.zip where minor is the last survey in
   the export.

"""
import argparse
import codecs
from collections import OrderedDict
from copy import deepcopy
import csv
import sys
from openpyxl.styles import Font # , Color, NamedStyle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from assign_nums import num_dict

TO_COMPARE = {'q13':  # likely to recommend
              'q16'   # age
              }
# qtext - the text name of the question, from row 2
# ans_dict - dictionary mapping the answer text to the count


class Qdata:

    def __init__(self, qnum):
        self.qnum = qnum
        self.startcol = q_dict[qnum]
        # For example, if our question is q13, limitcol is q14's column number.
        self.limitcol = q_dict['q' + str(int(qnum[1:]) + 1)]
        self.qtext = question_text_row[self.startcol]

        # For the major question, the values of the ans_dict are minor question
        # Qdata instances. For the minor questions, the values are integers
        # containing the counts of occurences.

        self.ans_dict = OrderedDict([(answer_text_row[n], 0)
                                    for n in range(self.startcol,
                                                   self.limitcol)])
        self.total = 0  # valid responses
        self.base = 0  # all responses


def trace(level, template, *args):
    if _args.verbose >= level:
        print(template.format(*args))


def oneans(row, qdminor):
    """
    Iterate over the possible minor answers and increment the counter when
    they exist.
    :param row:
    :param qdminor: the Qdata for the minor question that corresponds to the
                    current major question's current answer
    :return: 0 if at least 1 answer is found, 2 if no answers are found
    """
    error = 2  # 2 => error is in minor question
    for anscol in range(qdminor.startcol, qdminor.limitcol):
        if row[anscol]:
            anstext = answer_text_row[anscol]  # minor answer
            qdminor.ans_dict[anstext] += 1
            error = 0
            if not _args.multiple:
                break
    return error


def onerow(row, qdmajor, qdminor):
    """
    Iterate over the possible major answers. If an answer is present then
    iterate over its minor answers and increment each minor answer that is
    present.

    :param row:
    :param qdmajor: The current major question's Qdata
    :param qdminor: The current minor template Qdata
    :return:
    """
    qdmajor.base += 1
    error = 1  # 1 => error is in major question
    for anscol in range(qdmajor.startcol, qdmajor.limitcol):
        if row[anscol]:
            anstext = answer_text_row[anscol]  # Male / Female
            error = oneans(row, qdmajor.ans_dict[anstext])
            break
    if error and _args.verbose > 1:
        badquestion = ('', qdmajor.qnum, qdminor.qnum)[error]
        print(f'*****skipping row {qdmajor.base} respondent {row[0]}'
              f', no response to {badquestion}')


def make_major_qdata(major, infile):

    """

    :param major:  the question for the left column, a string like "q4"
    :param infile:
    :return: this question's populated major QData and the template minor
    Qdata which contains the minor question and answers and is more convenient
    to use than digging the minor info out of the major structure.
    """
    global q_dict, question_text_row, answer_text_row
    reader = csv.reader(infile)
    question_row = next(reader)  # has values like q1,,,,q2,,,q3,,etc.
    q_dict = num_dict(question_row)  # dict mapping q<minor> -> column index
    question_text_row = next(reader)
    answer_text_row = next(reader)
    minor = TO_COMPARE[major]
    qdmajor = Qdata(major)
    qdminor = Qdata(minor)  # minor qdata template
    '''
        For each major answer, create a minor qdata and store it as the value
        of the major's answer dictionary, overwriting the zero that was created
        by make_qdata.
    '''
    for ans in qdmajor.ans_dict:
        qdmajor.ans_dict[ans] = deepcopy(qdminor)  # each answer gets its own
    for row in reader:
        onerow(row, qdmajor, qdminor)
    return qdmajor, qdminor


def count_answers(major_qdata, minor_qdata):
    for majans in major_qdata.ans_dict:
        print(majans)
        minq = major_qdata.ans_dict[majans]
        print(minq.qtext)
        for ans in minq.ans_dict:
            print(ans, minq.ans_dict[ans])
            minq.total += minq.ans_dict[ans]
        print(minq.total)
        major_qdata.total += minq.total
    globalmin = minor_qdata.ans_dict
    # accumulate minor answers across all major answers
    for majans in major_qdata.ans_dict:  # iterate over major answers
        minq = major_qdata.ans_dict[majans]  # minor Qdata
        for minans in minq.ans_dict:  # iterate over minor answers
            globalmin[minans] += minq.ans_dict[minans]  # all minor answers


def create_xlsx(major_qdata, minor_qdata):
    j = major_qdata
    wb = Workbook()
    ws = wb.active
    title = f'{j.qnum.upper()}: {j.qtext.upper()}'
    a1 = ws.cell(row=1, column=1, value=title)
    a2 = ws.cell(row=2, column=1, value='BASE: ALL RESPONDENTS')
    a3 = ws.cell(row=5, column=1, value='BASE')
    b3 = ws.cell(row=5, column=2, value=j.base)
    b6 = ws.cell(row=6, column=2, value=1.0)  # 100%
    ws.cell(row=7, column=1, value='VALID RESPONSES')
    b7 = ws.cell(row=7, column=2, value=j.total)
    b8 = ws.cell(row=8, column=2, value=j.total / j.base)
    c3 = ws.cell(row=3, column=3, value=
                 f'{minor_qdata.qnum.upper()}: {minor_qdata.qtext.upper()}')
    # Populate the minor answer names
    coln = 2  # start with column 3
    for ans in minor_qdata.ans_dict:
        coln += 1
        ws.cell(row=4, column=coln, value=ans)
        # ws.cell(row=6, column=coln, value=1.0).style = 'Percent'
        ws.column_dimensions[get_column_letter(coln)].width = len(ans) * 1.20
    a1.font = Font(bold=True)
    a2.font = Font(bold=True)
    a3.font = Font(bold=True)
    b3.font = Font(bold=True)
    b7.font = Font(bold=True)
    c3.font = Font(bold=True)
    b6.style = 'Percent'
    b8.style = 'Percent'
    # Set the column width of the first column
    colw = 0
    for q in j.ans_dict:
        w = len(q)
        if w > colw:
            colw = w
    ws.column_dimensions['A'].width = colw * 1.10
    # Populate the minor answer totals
    globalans = minor_qdata.ans_dict
    col = 2  # 1st col = 3
    for minans in globalans:
        col += 1
        c = ws.cell(row=7, column=col, value=globalans[minans])
        c.font = Font(bold=True)
        c = ws.cell(row=8, column=col,
                    value=globalans[minans] / j.total)
        c.style = 'Percent'
    # For each major answer, populate the corresponding minor answers
    rownum = 7
    for majans in j.ans_dict:  # iterate over the major answers
        rownum += 3
        ws.cell(row=rownum, column=1, value=majans)
        minqd = j.ans_dict[majans]
        c = ws.cell(row=rownum, column=2, value=minqd.total)
        c.font = Font(bold=True)
        c = ws.cell(row=rownum + 1, column=2, value=minqd.total / j.total)
        c.style = 'Percent'
        col = 2
        for minans in minqd.ans_dict:
            col += 1
            c = ws.cell(row=rownum, column=col, value=minqd.ans_dict[minans])
            c.font = Font(bold=True)
            if minqd.total:
                c = ws.cell(row=rownum + 1, column=col,
                            value=minqd.ans_dict[minans] / minqd.total)
                c.style = 'Percent'
            else:
                c = ws.cell(row=rownum + 1, column=col, value='NaN')

    wb.save(sys.argv[2])


def main():
    for question in TO_COMPARE:
        with codecs.open(sys.argv[1], 'r', 'utf-8-sig') as infile:
            major_qdata, minor_qdata = make_major_qdata(question, infile)
            # print('major_qdata:')
            # print(major_qdata)
            print("Major question:", major_qdata.qtext)
            print("Minor question:", minor_qdata.qtext)
            count_answers(major_qdata, minor_qdata)
            create_xlsx(major_qdata, minor_qdata)
            print('Major Qdata total', major_qdata.total)


def getargs():
    parser = argparse.ArgumentParser(
        description='''

        ''')
    parser.add_argument('infile', help='''
         The CSV file that has been cleaned by extract_csv.sh''')
    parser.add_argument('outfile',
                        help='''output XLSX file.
        ''')
    parser.add_argument('-v', '--verbose', default=1, type=int, help='''
    ''')
    parser.add_argument('-m', '--multiple', action='store_true', help='''
    Specifies whether the minor question may have multiple (or no) answers. If
    not specified, exactly one minor answer must be selected.''')
    parser.add_argument('-j', '--major', help='''Not implemented yet.''')
    parser.add_argument('-minor', '--minor', help='''Not implemented yet.''')
    args = parser.parse_args()
    return args


if __name__ == '__main__':
    if sys.version_info.major < 3 or sys.version_info.minor < 6:
        raise ImportError('requires Python 3.6')
    q_dict, question_text_row, answer_text_row = None, None, None
    _args = getargs()
    main()


