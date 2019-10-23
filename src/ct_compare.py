# -*- coding: utf-8 -*-
"""
    ct_compare.py -- Compare the major values between two crosstab files.

    Input: the two crosstab files to compare.
    Output: An XLSX file with the compared result.
"""

import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os.path
import sys

CENTER = Alignment(horizontal='center')


def getyear(filename):
    # The filename is in format: crosstab_999_yyyy.xlsx
    tail = os.path.split(filename)[1]
    root = os.path.splitext(tail)[0]
    year = root[-4:]  # get yyyy
    print(f'year = {year}')
    return int(year)


def one_row(r, ws1, ws2, ws3):
    def format_cell():
        if not label:
            cell.style = 'Percent'
        elif label == 'MEAN VALUE':
            cell.number_format = '#0.00'

    label = ws1.cell(row=r, column=1).value
    ws3.cell(row=r, column=1, value=ws1.cell(row=r, column=1).value)
    cell = ws3.cell(row=r, column=2, value=ws1.cell(row=r, column=2).value)
    format_cell()
    cell = ws3.cell(row=r, column=3, value=ws2.cell(row=r, column=2).value)
    format_cell()


def main():
    wb1 = load_workbook(_args.in1)
    wb2 = load_workbook(_args.in2)
    y1 = getyear(_args.in1)
    y2 = getyear(_args.in2)
    wb3 = Workbook()
    del wb3[wb3.sheetnames[0]]
    for sheetname in wb1.sheetnames:
        ws1 = wb1[sheetname]
        ws2 = wb2[sheetname]
        ws3 = wb3.create_sheet(sheetname)
        ws3['A1'].value = ws1['A1'].value  # Title
        ws3['A2'].value = ws1['A2'].value
        ws3['B3'].value = y1
        ws3['B3'].alignment = CENTER
        ws3['C3'].value = y2
        ws3['C3'].alignment = CENTER

        for r in range(5, ws1.max_row):
            one_row(r, ws1, ws2, ws3)
        ws3.column_dimensions['A'].width = ws1.column_dimensions['A'].width
    wb3.save(_args.outfile)


def getargs():
    parser = argparse.ArgumentParser(
        description='''
        Create an XLSX file comparing the major values from two crosstabs
        files. The names of the input files are expected to be in the format
        "crosstab_nnn_xxxx". The column labels will be xxxx.
        ''')
    parser.add_argument('in1', help='''
         The first XLSX file created by crosstabs3.py''')
    parser.add_argument('in2', help='''
         The second XLSX file created by crosstabs3.py''')
    parser.add_argument('-o', '--outfile',
                        help='''output XLSX file.
        ''')
    parser.add_argument('-v', '--verbose', default=1, type=int, help='''
    Modify verbosity.
    ''')
    args = parser.parse_args()
    return args


if __name__ == '__main__':
    if sys.version_info.major < 3 or sys.version_info.minor < 6:
        raise ImportError('requires Python 3.6')
    _args = getargs()
    main()
    print('End ct_compare.')
