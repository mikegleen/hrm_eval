"""

crosstabs4 - based on crosstabs3 plus additional columns for answers broken
             down by year.

Input is a CSV file produced by extract_csv.sh. The creation method is:
1. Click on "Analyze Results".
2. Click on "SAVE AS".
3. Click on "Export file"
4. Select the dropdown "All response data".
5. Select the option ".CSV" (the leftmost option).
6. In "Data View", select "Original View".
7. In "Columns", select "Expanded".
8. In "Cells", select "Actual Answer Text".
9. Rename the file in the format yyyy-mm-dd_n.zip where n is the last survey in
   the export.
An input file might be:
    ~/Downloads/hrm/evaluation/data_exports/2017-11-05/cleaned/126.csv
or (via the symbolic link)
    ~/pyprj/hrm/evaluation/exports/2017-11-05/cleaned/126.csv

The cleaned file is processed before input to crosstabs3:
    split.py     -- splits question 9 into its subquestions 9.01, 9.02, etc.
    aggregate.py -- consolidates some answers. For example question 13 has
    answers of 'Very unlikely', 'Unlikely', 'Neither likely nor unlikely',
    'Likely', 'Very likely'. These are consolidated to 'Not very' or 'Likely'.
"""
import argparse
import codecs
from collections import OrderedDict, defaultdict
import csv
import sys
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import Union

from assign_nums import num_dict
from config import SKIPCOLS
from config import CROSSTAB_TITLES as TITLES
# MAJOR_ and MINOR_ QUESTIONS are lists of strings like ['Q1', 'Q2', ...]
from config import MAJOR_QUESTIONS, MINOR_QUESTIONS, SANITY_QUESTION
#
# Constants for sheet creation:
# The row to insert minor titles and answer names
MINOR_NUMBER_ROW = 3
MINOR_QUESTION_NAME_ROW = 4
MINOR_ANSWER_NAME_ROW = 5
BASE_ROW = 6
VALID_RESPONSES_ROW = 8
# The row number where we initialize the loop that inserts row counts.
MINOR_COUNT_START = 12
MINOR_COUNT_INCREMENT = 3

LEFT_BORDER = Border(left=Side(border_style='thin'))
CENTER = Alignment(horizontal='center')
BOLD = Font(bold=True)
WRAP = Alignment(wrapText=True)
MIN_COL_WIDTH = 6.0
MAX_COL_WIDTH = 14.0

# TO_COMPARE: A dict for each major question the value is a list of minor
# questions excluding the current major question.
TO_COMPARE = {major: [minor for minor in MINOR_QUESTIONS if minor != major]
              for major in MAJOR_QUESTIONS}


class Qdata:

    def __init__(self, qnum: str):
        self.qnum = qnum
        self.startcol = q_dict[qnum]
        trace(3, 'qnum {}, startcol: {}', qnum, self.startcol)
        # For example, if our question is q13, limitcol is q14's column number.
        # An exception to this rule is when a question has been split, in
        # which case, for example, q9 is replaced by q9.01 ... q9.16. So we have
        # to use a roundabout method to get the "next" question's columm.
        # Note that there is a dummy entry at the end of q_dict which allows
        # the last legitimate qnum to compute a limitcol.
        qlist = list(q_dict)
        limitix = qlist[qlist.index(qnum) + 1]
        self.limitcol = q_dict[limitix]
        trace(3, '  limitix: {}, limitcol: {}', limitix, self.limitcol)
        # qtext - the text name of the question, from row 2
        self.qtext = question_text_row[self.startcol]

        # For the major question, the values of the ans_dict are minor question
        # Qdata instances. Not used for the minor questions. Will be
        # initialized by make_major_qdata().
        self.ans_dict: Union[int, dict] = OrderedDict(
            [(answer_text_row[n], 0)
             for n in range(self.startcol, self.limitcol)])
        count_list = [(answer_text_row[n], 0) for n in range(self.startcol,
                                                             self.limitcol)]
        by_yr_list = [(t[0], defaultdict(int)) for t in count_list]

        # ans_count - dict mapping the answer text to the count
        self.ans_count = dict(count_list)

        # year_answers - dict mapping the answer text to a dict keyed by year
        self.year_answers = dict(by_yr_list)

        # minor_totals - a dict with key minor question # and value the
        # column totals for each of the answers of the minor question
        self.minor_totals = None  # will be populated for major question only

        # value_totals - a dict with key minor question # and value the
        # answer indices. For example, if the first question is given the
        # value 1 and so on, then value_totals will contain the sum.  This is
        # used to compute a (highly dubious) mean value for the minor question.
        self.value_totals = None
        self.year_totals = None
        self.year_value_totals = None
        self.year_base = None

        # Define a set of the valid years so that even if a row has no values
        # in a year, that column will still be counted.
        self.yearset = set()
        self.total = 0  # valid responses
        self.base = 0  # all responses


def trace(level, template, *args):
    if _args.verbose >= level:
        print(template.format(*args))


def validate_minor(row, qdminor):
    """
    Iterate over the possible minor answers and return None if they exist.
    :param row: the row from the CSV file.
    :param qdminor: the Qdata for the minor question that corresponds to the
                    current major question's current answer
    :return: None if an answer is found, otherwise the question number (Qnn).
    """
    error = qdminor.qnum

    for anscol in range(qdminor.startcol, qdminor.limitcol):
        if row[anscol]:
            error = None
            break
    return error


def validate_one_row(row, qdmajor):
    error = qdmajor.qnum  # error is in major question
    for anscol in range(qdmajor.startcol, qdmajor.limitcol):
        if row[anscol]:
            error = None  # found a major answer
            anstext = answer_text_row[anscol]  # Male / Female
            minordict = qdmajor.ans_dict[anstext]
            # For each of the minor questions, if any have no answers,
            # the row is invalid.
            for minorqdata in minordict.values():
                minorerror = validate_minor(row, minorqdata)
                if minorerror:
                    error = minorerror
                    break
    return error


def count_minor_question(row, qdminor):
    """
    Iterate over the possible minor answers and increment the counter when
    they exist.
    :param row:
    :param qdminor: the Qdata for the minor question that corresponds to the
                    current major question's current answer
    :return: None
    """
    for anscol in range(qdminor.startcol, qdminor.limitcol):
        if row[anscol]:
            trace(3, 'count_minor_question: minor: {}, col: {}', qdminor.qnum,
                  anscol)
            anstext = answer_text_row[anscol]  # minor answer
            qdminor.ans_count[anstext] += 1


def count_one_row(row, qdmajor: Qdata):
    """
    Iterate over the possible major answers. If an answer is present then
    iterate over its minor questions and increment each minor answer that is
    present.

    :param row:
    :param qdmajor: The current major question's Qdata
    :return: None
    """
    # row[2] is like '2017-12-08T19:42:01Z'
    ansyear = int(row[2][:4])
    if _args.year and ansyear != _args.year:
        return
    if _args.oldestyear and ansyear < _args.oldestyear:
        return
    qdmajor.yearset.add(ansyear)
    qdmajor.base += 1
    qdmajor.year_base[ansyear] += 1
    # First check that there is at least one answer to each minor question
    error = validate_one_row(row, qdmajor) if _args.complete else False
    if error:
        trace(2, '*****skipping row {} respondent {}, no response to {}',
              qdmajor.base, row[0], error)
        return

    found = False
    for anscol in range(qdmajor.startcol, qdmajor.limitcol):
        trace(3, 'count_one _row: major: {}, col: {}', qdmajor.qnum, anscol)
        if row[anscol]:
            found = True
            anstext = answer_text_row[anscol]  # Male / Female
            qdmajor.ans_count[anstext] += 1
            qdmajor.year_answers[anstext][ansyear] += 1
            qdmajor.total += 1
            minordict = qdmajor.ans_dict[anstext]
            for minorqdata in minordict.values():
                count_minor_question(row, minorqdata)
    if not found:
        trace(2, '*****skipping row {} respondent {}, no response to {}',
              qdmajor.base, row[0], qdmajor.qnum)


def make_major_qdata(major, infile):
    """
    :param major:  the question for the left column, a string like "q4"
    :param infile: a pass is made over the data for each major question.
    :return: this question's populated major QData and the template minor
    Qdata which contains the minor question and answers and will be used to
    accumulate totals.

    Also create the global "q_dict" which maps question number to column.
    """
    global q_dict, question_text_row, answer_text_row
    reader = csv.reader(infile)
    question_row = next(reader)  # has values like q1,,,,q2,,,q3,,etc.
    if SANITY_QUESTION.lower() not in question_row:
        print('Invalid CSV file. Maybe not the output of aggregate->split.')
        sys.exit(1)
    # Create a dict mapping Q<minor> -> column index
    question_row.append('qx')  # dummy column at end for Qdata constructor
    # map question number (like 'Q4') to column
    q_dict = num_dict(question_row, _args.skipcols)
    question_text_row = next(reader)
    answer_text_row = next(reader)
    minortuple: list = TO_COMPARE[major]
    qdmajor = Qdata(major)
    '''
        For each major answer, create a dict of minor qdata and store it as the
        value of the major's answer dictionary.
    '''
    for ans in qdmajor.ans_dict:
        qdmajor.ans_dict[ans] = {minor: Qdata(minor) for minor in minortuple}
    # Create a dictionary of dictionaries where the outer dictionary is keyed
    # by the minor questions and the inner dictionaries are each keyed by the
    # respective minor questions answers.
    qdmajor.minor_totals = {}
    qdmajor.value_totals = {}
    qdmajor.year_totals = defaultdict(int)
    qdmajor.year_value_totals = defaultdict(int)
    qdmajor.year_base = defaultdict(int)
    for minor in minortuple:
        minorqd = Qdata(minor)
        qdmajor.minor_totals[minor] = {ans: 0 for ans in minorqd.ans_dict}
        qdmajor.value_totals[minor] = {ans: 0 for ans in minorqd.ans_dict}
    for row in reader:
        count_one_row(row, qdmajor)
    return qdmajor


def count_answers(major_qdata: Qdata):
    """
    Called once for each major question.

    :param major_qdata:
    :return: None
    """
    for ix, majans in enumerate(major_qdata.ans_dict, start=1):
        major_answer_by_year = major_qdata.year_answers[majans]
        for yr in major_answer_by_year:
            major_qdata.year_totals[yr] += major_answer_by_year[yr]
            major_qdata.year_value_totals[yr] += major_answer_by_year[yr] * ix
        trace(3, 'in count_answers, majans = {}', majans)
        minor_qdata_dict = major_qdata.ans_dict[majans]
        for minq in minor_qdata_dict.values():
            trace(3, 'minor question: {}', minq.qtext)
            for ans in minq.ans_count:
                trace(3, r'ans: "{}" ({})', ans, minq.ans_count[ans])
                minq.total += minq.ans_count[ans]
            trace(3, 'in count_answers, minq.total = {}', minq.total)

    minq_tuple = TO_COMPARE[major_qdata.qnum]
    for minq in minq_tuple:
        # anstotal: For this major question and this minor question, anstotal
        # is a dictionary where the key is the minor answer and the value is
        # the sum of the minor answer counts.
        anstotal = major_qdata.minor_totals[minq]
        # valuetotal will be used to compute the mean value
        valuetotal = major_qdata.value_totals[minq]
        # accumulate minor answers across all major answers
        for ix, majans in enumerate(major_qdata.ans_dict, start=1):
            minor_qdata_dict = major_qdata.ans_dict[majans]  # minor Qdata
            minqdata = minor_qdata_dict[minq]
            for minans in minqdata.ans_count:  # iterate over minor answers
                anstotal[minans] += minqdata.ans_count[minans]
                valuetotal[minans] += minqdata.ans_count[minans] * ix


def setvalue(worksheet, row, column, value, total):
    """
    Insert a count and immediately below it insert the percent of the given
    total.
    """
    # print(f"row={row}, col={column}, value={value}, total={total}")
    cell = worksheet.cell(row=row, column=column, value=value)
    cell.font = Font(bold=True)
    if total is None:
        return  # Don't put percent
    if total:
        cell = worksheet.cell(row=row + 1, column=column, value=value / total)
        cell.style = 'Percent'
    else:  # avoid divide by zero
        cell = worksheet.cell(row=row + 1, column=column, value='-')
        cell.alignment = CENTER


def setmean(worksheet, row, column, value, count):
    try:
        meanvalue = float(value) / float(count)
    except ZeroDivisionError:
        meanvalue = None
    if meanvalue is None:
        cell = worksheet.cell(row=row, column=column, value='-')
        cell.alignment = CENTER
    else:
        cell = worksheet.cell(row=row, column=column, value=meanvalue)
        cell.number_format = '#0.00'
    cell.font = BOLD


def one_minor(ws, major_qdata, minor_qnum, startcol):
    """
    Create the cells in this major question's worksheet for one minor question.

    :param ws: the current worksheet that we're creating
    :param major_qdata:
    :param minor_qnum: string in the form 'Q13'
    :param startcol: column in the worksheet to start inserting this minor
                     question and its answers
    :return: the last used minor Qdata which will be used by the caller to
             extract the start and end column values.
             Note that all the Qdata instances for this minor question (one
             for each major answer) have the same start/end column values.
    """
    # put the total values in the "VALID RESPONSES" row.
    col = startcol - 1
    minor_qdata = None  # avoid warnings
    row = 0  # avoid warnings
    # Iterate over the answers for this minor question.
    for minans, mincount in major_qdata.minor_totals[minor_qnum].items():
        col += 1
        cell = ws.cell(row=MINOR_ANSWER_NAME_ROW, column=col, value=minans)
        cell.alignment = Alignment(wrapText=True)
        width = len(minans) * 1.10
        width = MIN_COL_WIDTH if width < MIN_COL_WIDTH else width
        width = MAX_COL_WIDTH if width > MAX_COL_WIDTH else width
        ws.column_dimensions[get_column_letter(col)].width = width
        row = VALID_RESPONSES_ROW
        setvalue(ws, row, col, mincount, major_qdata.total)
        minortotal = major_qdata.minor_totals[minor_qnum][minans]
        valuetotal = major_qdata.value_totals[minor_qnum][minans]
        # Iterate over the major answers
        row = MINOR_COUNT_START
        for majans, minor_qdata_dict in major_qdata.ans_dict.items():
            minor_qdata = minor_qdata_dict[minor_qnum]
            setvalue(ws, row, col, minor_qdata.ans_count[minans], minortotal)
            row += MINOR_COUNT_INCREMENT
        setmean(ws, row, col, valuetotal, minortotal)
    for r in range(3, row + 1):
        ws.cell(row=r, column=startcol).border = LEFT_BORDER
    if minor_qnum in TITLES:
        txt = TITLES[minor_qnum]
    else:
        txt = minor_qdata.qtext
    cell = ws.cell(row=MINOR_NUMBER_ROW, column=startcol, value=minor_qnum)
    cell.font = BOLD
    cell = ws.cell(row=MINOR_QUESTION_NAME_ROW, column=startcol,
                   value=txt.upper())
    cell.font = BOLD
    return minor_qdata


def one_year(ws, major_qdata, year, col):
    width = MIN_COL_WIDTH
    ws.column_dimensions[get_column_letter(col)].width = width
    cell = ws.cell(row=MINOR_NUMBER_ROW, column=col, value=str(year))
    cell.font = BOLD
    cell.alignment = CENTER
    # print(major_qdata.year_base)
    # print(year, major_qdata.year_base[year])
    setvalue(ws, BASE_ROW, col, major_qdata.year_base[year], major_qdata.base)
    row = VALID_RESPONSES_ROW
    total = major_qdata.year_totals[year]
    setvalue(ws, row, col, total, major_qdata.year_base[year])

    row = MINOR_COUNT_START
    for ans in major_qdata.year_answers:
        value = major_qdata.year_answers[ans][year]
        setvalue(ws, row=row, column=col, value=value, total=total)
        row += MINOR_COUNT_INCREMENT
    setmean(ws, row, col, major_qdata.year_value_totals[year],
            major_qdata.year_totals[year])


def one_sheet(major_qdata):
    """
    Produce a sheet like:
1   |Q13: HOW LIKELY ARE YOU TO RECOMMEND...
2   |BASE: ALL RESPONDENTS
3   |               |          |Q16: YOUR AGE
4   |               |          |Under 55 |55 or over|
5   |BASE           |       159|         |          |
6   |               |      100%|         |          |
    |VALID RESPONSES|       136|       39|        97|
    |               |       86%|      29%|       71%|
    |               |          |         |          |
    |Not very       |        19|       18|        11|
    |               |       14%|      21%|       11%|
    |               |          |         |          |
    |Likely         |       117|       31|        86|
    |               |       86%|      79%|       89%|

    :param major_qdata:
    :return: None. The workbook is updated.
    """
    major_qdata = major_qdata
    ws = workbook.create_sheet(major_qdata.qnum)
    title = f'{major_qdata.qnum.upper()}: {major_qdata.qtext.upper()}'
    a1 = ws.cell(row=1, column=1, value=title)
    a2 = ws.cell(row=2, column=1, value='BASE: ALL RESPONDENTS')
    a6 = ws.cell(row=6, column=1, value='BASE')
    b6 = ws.cell(row=6, column=2, value=major_qdata.base)
    b7 = ws.cell(row=7, column=2, value=1.0)  # 100%
    ws.cell(row=8, column=1, value='VALID RESPONSES')
    b8 = ws.cell(row=8, column=2, value=major_qdata.total)
    b9 = ws.cell(row=9, column=2, value=major_qdata.total / major_qdata.base)
    a1.font = BOLD
    a2.font = BOLD
    a6.font = BOLD
    b6.font = BOLD
    b8.font = BOLD
    b7.style = 'Percent'
    b9.style = 'Percent'
    # Set the column width of the first column
    colw = 22
    for q in major_qdata.ans_dict:
        w = len(q)
        if w > colw:
            colw = w
    ws.column_dimensions['A'].width = colw  # * 1.10

    # Insert the major answers and the response totals.
    rownum = MINOR_COUNT_START
    index = 0
    value_total = 0
    for majans in major_qdata.ans_dict:  # iterate over the major answers
        index += 1
        ws.cell(row=rownum, column=1, value=f'({index}) ' + majans)
        ans_total = major_qdata.ans_count[majans]
        setvalue(ws, rownum, 2, ans_total, major_qdata.total)
        value_total += ans_total * index
        rownum += MINOR_COUNT_INCREMENT
    ws.cell(row=rownum, column=1, value='MEAN VALUE').font = BOLD
    try:
        mean_value = float(value_total) / float(major_qdata.total)
        cell = ws.cell(row=rownum, column=2, value=mean_value)
        cell.number_format = '#0.00'
    except ZeroDivisionError:
        cell = ws.cell(row=rownum, column=2, value='-')
        cell.alignment = CENTER
    cell.font = BOLD
    comment = ('The mean value must be used with caution. The values are'
               ' arbitrarily assiged with the first answer in a column given a'
               ' value of 1 and so on.')
    ws.cell(row=rownum + 2, column=3, value=comment)

    # Iterate over the minor questions, inserting the minor answers.
    coln = 3
    for year in sorted(major_qdata.yearset):
        one_year(ws, major_qdata, year, coln)
        coln += 1
    for minor in major_qdata.minor_totals:
        minor_qdata = one_minor(ws, major_qdata, minor, coln)
        minorlen = minor_qdata.limitcol - minor_qdata.startcol
        coln += minorlen
    ws.freeze_panes = 'C6'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def main():
    global workbook
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]  # remove the default sheet
    for question in MAJOR_QUESTIONS:
        with codecs.open(_args.infile, 'r', 'utf-8-sig') as infile:
            trace(2, "Major question: {}", question)
            major_qdata = make_major_qdata(question, infile)
            count_answers(major_qdata)
            one_sheet(major_qdata)
            text: str = major_qdata.qtext
            if len(text) > 50:
                text = text[:50] + '...'
            trace(1, 'Major question {}: "{}" total {}', major_qdata.qnum,
                  text, major_qdata.total)
    workbook.save(_args.outfile)


def getargs():
    parser = argparse.ArgumentParser(description='''
        Create a crosstabs xlsx file.
        ''')
    parser.add_argument('infile', help='''
         The CSV file that has been processed by extract_csv.sh. The
         full tool chain clean->aggregate->split is required to put the
         input file into the correct format.''')
    parser.add_argument('outfile',
                        help='''output XLSX file.
        ''')
    parser.add_argument('-c', '--complete', action='store_true', help='''
                        If specified, require that each minor question has
                        at least one answer otherwise the row is rejected.''')
    parser.add_argument('-o', '--oldestyear', type=int, default=0,
                        help='''
                        If specified, ignore responses before this year.
                        ''')
    parser.add_argument('-s', '--skipcols', type=int, default=SKIPCOLS,
                        help=f'''Number of columns to ignore before extracting
                        the column numbers for defined questions. Default is
                        {SKIPCOLS}.''')
    parser.add_argument('-v', '--verbose', default=1, type=int, help='''
    Modify verbosity.
    ''')
    parser.add_argument('-y', '--year', type=int, default=0,
                        help='''
                        If specified, only process responses from this year.
                        ''')
    args = parser.parse_args()
    return args


if __name__ == '__main__':
    assert sys.version_info >= (3, 6)
    _args = getargs()
    # avoid global variable warning
    workbook: Workbook | None = None
    q_dict, question_text_row, answer_text_row = {}, [], []
    main()
    print('End crosstabs4.')
