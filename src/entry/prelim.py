"""

"""
import argparse
from collections import defaultdict
import codecs
import datetime
import csv
import os.path
import sys
from openpyxl import Workbook, load_workbook


def getargs():
    parser = argparse.ArgumentParser(
        description='''
        Clean the data from the Excel spreadsheet.
        Count the ways visitors have heard about the museum.
        ''')
    parser.add_argument('infile', help='''
         The input XLSX file containing multiple tabs.''')
    parser.add_argument('outfile',
                        help='''output CSV file.
        ''')
    parser.add_argument('howheard',
                        help='''output CSV file for "how heard" tally.
        ''')
    parser.add_argument('-v', '--verbose', default=1, type=int, help='''
    Modify verbosity.
    ''')
    args = parser.parse_args()
    return args


def main():
    oldwb = load_workbook(_args.infile, read_only=True)

    # Count the ways visitors have heard about the museum

    how_heard: defaultdict = defaultdict(int)
    sheet = oldwb['Data']
    n = nw = 0

    for row in sheet.iter_rows(max_row=4760, max_col=12):
        n += 1
        # Get rid of Excel formulas
        row = ['' if not r.value or type(r.value) is str and
               r.value.startswith('=') else r.value for r in row]
        if _args.verbose > 1:
            print(n, row)
        try:
            # Convert the datetime into a date string because the default
            # is to emit the date and time.
            row[2] = row[2].strftime('%Y-%m-%d')
        except AttributeError:
            pass  # Isn't a datetime object so just pass on whatever it is
        if any(row):
            nw += 1
            writer.writerow(row)

        if len(row) > 5 and row[5]:
            how_heard[row[5].lower().strip()] += 1
    # print(how_heard)
    for how in sorted(how_heard, key=how_heard.get, reverse=True):
        print(f'{how},{how_heard[how]}', file=howheard)
    print(f'{n} rows read from {_args.infile}')
    print(f'{nw} rows written to {_args.outfile}')
    print(f'{len(how_heard)} rows written to {_args.howheard}')


if __name__ == '__main__':
    assert sys.version_info >= (3, 6)
    _args = getargs()
    if _args.verbose > 1:
        print(f'verbosity: {_args.verbose}')
    outfile = codecs.open(_args.outfile, 'w', 'utf-8-sig')
    howheard = codecs.open(_args.howheard, 'w', 'utf-8-sig')

    writer = csv.writer(outfile)
    main()
    print(f'End {os.path.splitext(os.path.basename(sys.argv[0]))[0]}.')
