"""

"""
import argparse
from collections import defaultdict
import codecs
import csv
import sys
from openpyxl import Workbook, load_workbook

HDG_DATE = 'DATE'
HDG_VISITORS = 'No. of visitors in Party'
HDG_FIRSTVISIT = 'First Visit? Y/N'
HDG_WHEREFROM = 'Where did they travel from'


def getargs():
    parser = argparse.ArgumentParser(
                                     description='''
        Clean the data from the Excel spreadsheet
        ''')
    parser.add_argument('infile', help='''
         The input XLSX file containing multiple tabs.''')
    parser.add_argument('-o', '--outfile',
                        help='''output CSV file.
        ''')
    parser.add_argument('--dryrun', action='store_true', help='''
        Do not write to the output file.''')
    parser.add_argument('-f', '--function', type=int, required=True, help='''
1. List heading rows.
2. Find bad dates.
3. Merge sheets to CSV file including the date and where-from columns.
    ''')
    parser.add_argument('-v', '--verbose', default=1, type=int, help='''
    Modify verbosity.
    ''')
    args = parser.parse_args()
    return args


def list_hdg_rows(wb):  # function 1
    sheets = wb.worksheets
    for sheet in sheets[3:]:
        row = sheet[1]
        # print([r.value for r in row], sheet.title)
        if not _args.dryrun:
            writer.writerow([sheet.title] + [r.value for r in row])
        # print(sheet.title, type(sheet.title))


def list_bad_dates(wb):  # function 2
    sheets = wb.worksheets
    for sheet in sheets:
        irows = sheet.iter_rows(max_row=4760, max_col=12)
        next(irows)
        nrow = 1
        for row in irows:
            nrow += 1
            row = [r.value for r in row[:5]]
            try:
                # Convert the datetime into a date string because the default
                # is to emit the date and time.
                row[2] = row[2].strftime('%Y-%m-%d')
            except AttributeError:
                if any(row):
                    print(sheet.title, nrow, [r for r in row])
            # writer.writerow([sheet.title] + [r.value for r in row])


def merge_sheets(wb):  # function 3
    sheets = list(wb.worksheets)[3:]
    row = 'date visitors first_visit wherefrom'.split()
    if not _args.dryrun:
        writer.writerow(row)  # write header
    for sheet in sheets:
        # if sheet.title != 'January 2020':
        #     continue
        irows = sheet.iter_rows(max_row=1000, max_col=12)
        titlerow = [r.value for r in next(irows)]
        # print(titlerow)
        nrow = 1
        if (date_col := titlerow.index(HDG_DATE)) < 0:
            print(f'Cannot find DATE in sheet {sheet.title}.')
            continue
        if (visitors_col := titlerow.index(HDG_VISITORS)) < 0:
            print(f'Cannot find HDG_VISITORS in sheet {sheet.title}.')
            continue
        if (first_col := titlerow.index(HDG_FIRSTVISIT)) < 0:
            print(f'Cannot find FIRSTVISIT in sheet {sheet.title}.')
            continue
        if (from_col := titlerow.index(HDG_WHEREFROM)) < 0:
            print(f'Cannot find WHEREFROM in sheet {sheet.title}.')
            continue
        # print(f'{date_col=}, {from_col=}')
        for row in irows:
            row = [r.value for r in row]
            try:
                # Convert the datetime into a date string because the default
                # is to emit the date and time.
                if not any(row):
                    # print(sheet.title, nrow, ' skipping.')
                    continue
                # print(row)
                rowdate = row[date_col].strftime('%Y-%m-%d')
                nrow += 1
                if not _args.dryrun:
                    writer.writerow([rowdate, row[visitors_col],
                                     row[first_col], row[from_col]])
            except AttributeError as err:
                print(f'Attribute error: {err}')
                if any(row):
                    print('Bad date: ', sheet.title, nrow, [r for r in row])
        print(f'sheet {sheet.title} {nrow=}')


def main():
    wb = load_workbook(_args.infile)
    if _args.function == 1:
        list_hdg_rows(wb)
    elif _args.function == 2:
        list_bad_dates(wb)
    elif _args.function == 3:
        merge_sheets(wb)

    # for row in sheet.values:
    #     n += 1
    #     print(n, row)
    #     writer.writerow(row)
    #     if len(row) > 5 and row[5]:
    #         how_heard[row[5].lower().strip()] += 1
    # print(how_heard)
    # for how in sorted(how_heard, key=how_heard.get, reverse=True):
    #     print(how, how_heard[how])


if __name__ == '__main__':
    assert sys.version_info >= (3, 6)
    writer = None
    _args = getargs()
    if _args.verbose > 1:
        print(f'verbosity: {_args.verbose}')
    print(f'Input: {_args.infile}')
    outfile = codecs.open(_args.outfile, 'w', 'utf-8-sig')
    if not _args.dryrun:
        writer = csv.writer(outfile)
    main()
